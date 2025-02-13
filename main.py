import socket
import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


# --------------------------------------------------
# Utility Functions
# --------------------------------------------------
def get_clean_hostname():
    hostname = socket.gethostname()
    if hostname.endswith('.local'):
        hostname = hostname.replace('.local', '')
    return hostname

def get_base_output_path():
    if os.name == 'nt':
        return r'C:\Users\josemaria\Downloads'
    else:
        hostname = get_clean_hostname()
        if hostname == 'JM-MBP':
            return '/Users/j.m./Downloads'
        elif hostname == 'JM-MS':
            return '/Users/jm/Downloads'
        return None

def get_base_path(warehouse):
    """
    Returns the network base path for a given warehouse/folder,
    considering both Windows and macOS.

    Parameters:
      - warehouse (str): Should be one of 'OPL', 'E', or 'MOBU'
    """
    if os.name == 'nt':  # Windows paths
        if warehouse.upper() == 'OPL':
            return r'\\192.168.10.18\Bodega General\LIVSMART\BODEGA OPL\STOCK ACTUALIZADO - OPL'
        elif warehouse.upper() == 'E':
            return r'\\192.168.10.18\Bodega General\LIVSMART\BODEGA E\STOCK ACTUALIZADO - BODEGA E'
        elif warehouse.upper() == 'MOBU':
            return r'\\192.168.10.18\Bodega General\LIVSMART\BODEGAS MOBU\STOCK ACTUALIZADO - MOBU'
    else:  # macOS paths
        if warehouse.upper() == 'OPL':
            return '/Volumes/Bodega General/LIVSMART/BODEGA OPL/STOCK ACTUALIZADO - OPL/'
        elif warehouse.upper() == 'E':
            return '/Volumes/Bodega General/LIVSMART/BODEGA E/STOCK ACTUALIZADO - BODEGA E/'
        elif warehouse.upper() == 'MOBU':
            return '/Volumes/Bodega General/LIVSMART/BODEGAS MOBU/STOCK ACTUALIZADO - MOBU/'
    return None

def read_excel_file(file_path):
    """
    Attempts to read an Excel file using pandas (with openpyxl as the engine).

    Parameters:
      - file_path (str): Full path to the Excel file.

    Returns:
      - A DataFrame if successful; otherwise, None.
    """
    if os.path.exists(file_path):
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            print(f"Successfully loaded: {file_path}")
            return df
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return None
    else:
        print(f"File not found: {file_path}")
        return None

def list_directory_contents(warehouse):
    """
    Debug function: Lists all files in the base path for a given warehouse.
    """
    base_path = get_base_path(warehouse)
    if base_path is None:
        print(f"Could not determine base path for warehouse: {warehouse}")
        return
    print(f"\n--- Listing contents for warehouse {warehouse} ---")
    try:
        files = os.listdir(base_path)
        for f in files:
            print(f)
    except Exception as e:
        print(f"Error listing directory {base_path}: {e}")

# --------------------------------------------------
# Inventory Processing Function
# --------------------------------------------------
def process_inventory(inventory_type):
    """
    For a given inventory type, check all three warehouse folders and load the files that exist.

    Parameters:
      - inventory_type (str): Should be one of 'LATA', 'PREFORMA', or 'PT'

    Returns:
      - A dictionary of {warehouse: DataFrame} for each successfully read file.
    """
    inventory_files = {
        "LATA": {
            "OPL": "INVENTARIO LATA VACIA 2025 BODOPL.xlsx",
            "E": "INVENTARIO LATA VACIA 2025 BODE.xlsx",
            "MOBU": "INVENTARIO LATA VACIA 2025 MOBU.xlsx"
        },
        "PREFORMA": {
            "OPL": "INVENTARIO DE PREFORMA 2025 BODOPL.xlsx",
            "E": "INVENTARIO DE PREFORMA 2025 BODE.xlsx",
            "MOBU": "INVENTARIO DE PREFORMA 2025 BODMOBU.xlsx"
        },
        "PT": {
            "OPL": "INVENTARIO DE PT.XLSX",
            "E": "INVENTARIO DE PT.XLSX",
            "MOBU": "INVENTARIO DE PT.XLSX"
        }
    }

    results = {}
    warehouses = ["OPL", "E", "MOBU"]

    if inventory_type not in inventory_files:
        print(f"Unknown inventory type: {inventory_type}")
        return results

    for warehouse in warehouses:
        base_path = get_base_path(warehouse)
        if base_path is None:
            print(f"Could not determine base path for warehouse: {warehouse}")
            continue

        # Debug: List directory contents to verify available files
        list_directory_contents(warehouse)

        file_name = inventory_files[inventory_type].get(warehouse)
        if not file_name:
            print(f"No filename defined for {inventory_type} in warehouse {warehouse}.")
            continue

        file_path = os.path.join(base_path, file_name)
        print(f"\nLooking for file: {file_path}")
        df = read_excel_file(file_path)
        if df is not None:
            results[warehouse] = df
        else:
            print(f"Skipping {warehouse} for {inventory_type} as file could not be loaded.")

    return results

# --------------------------------------------------
# Merge Inventories Function
# --------------------------------------------------
def merge_inventories(inventory_dict):
    """
    For all loaded inventory DataFrames, add a new column 'Bodega' to indicate
    the warehouse source, and then merge them into a single DataFrame.

    Parameters:
      - inventory_dict (dict): Dictionary with keys as warehouse names and values as DataFrames.

    Returns:
      - A merged DataFrame with an additional column 'Bodega'
    """
    merged_list = []
    for warehouse, df in inventory_dict.items():
        if warehouse.upper() == "OPL":
            df["Bodega"] = "Bodega OPL"
        elif warehouse.upper() == "E":
            df["Bodega"] = "Bodega E"
        elif warehouse.upper() == "MOBU":
            df["Bodega"] = "Bodegas MOBU"
        else:
            df["Bodega"] = warehouse
        merged_list.append(df)

    if merged_list:
        merged_df = pd.concat(merged_list, ignore_index=True)
        return merged_df
    else:
        return None

# --------------------------------------------------
# Excel Formatting Function
# --------------------------------------------------
def format_excel_file(file_path):
    """
    Applies formatting to the Excel document:
      - Makes header row bold with a minimalist font.
      - Applies a minimalist font to all cells.
      - Adjusts column widths to fit the content.
      - Colors the 'Bodega' column cells based on their value.
      - Ensures 'CODIGO WMS' column contains only string values.

    Parameters:
      - file_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Define fonts
    header_font = Font(bold=True, name="Calibri")
    minimal_font = Font(name="Calibri")

    # Format header row (assumed to be the first row)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths based on maximum content length
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[column].width = max_length + 2

    # Apply minimalist font to all cells (then reapply header bold)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = minimal_font
    for cell in ws[1]:
        cell.font = header_font

    # Color the 'Bodega' column cells based on their value
    bodega_col = None
    codigo_wms_col = None

    for cell in ws[1]:
        if cell.value == "Bodega":
            bodega_col = cell.column_letter
        elif cell.value == "CODIGO WMS":
            codigo_wms_col = cell.column_letter

    if bodega_col is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{bodega_col}{row}"]
            if cell.value == "Bodega OPL":
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            elif cell.value == "Bodega E":
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif cell.value == "Bodegas MOBU":
                cell.fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")

    # Ensure 'CODIGO WMS' column contains only string values
    if codigo_wms_col is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{codigo_wms_col}{row}"]
            if cell.value is not None and not isinstance(cell.value, str):
                cell.value = str(cell.value)

    wb.save(file_path)


# --------------------------------------------------
# Main Function: Prompt User, Process, Merge & Format Inventories
# --------------------------------------------------
def main():
    print("=== Inventory Processing System ===")
    print("Presione el numero del producto del cual desea extraer el inventario:")
    print("\nLata: (1), Preforma: (2), PT: (3)")

    choice = input("Producto? ")

    mapping = {"1": "LATA", "2": "PREFORMA", "3": "PT"}
    inventory_type = mapping.get(choice.strip())
    if not inventory_type:
        print("Opción inválida. Saliendo.")
        return

    print(f"\n--- Processing {inventory_type} Inventory ---")
    results = process_inventory(inventory_type)

    print("\nInventory processing complete.")
    if results:
        print(f"\nProcessed {inventory_type} inventory from:")
        for wh in results:
            print(f"  - {wh}")

        # Merge inventories
        print("\n--- Merging Inventories ---")
        merged_inventory = merge_inventories(results)
        if merged_inventory is not None:
            print("Merged inventory preview:")
            print(merged_inventory.head())

            # Save as an Excel document with date-stamped filename
            output_path = get_base_output_path()
            if output_path:
                today_str = date.today().strftime("%Y%m%d")
                output_file = os.path.join(output_path, f"merged_inventory_{inventory_type}_{today_str}.xlsx")
                merged_inventory.to_excel(output_file, index=False, engine='openpyxl')
                print(f"Merged inventory saved to: {output_file}")

                # Apply formatting to the saved Excel file
                format_excel_file(output_file)
                print("Excel formatting applied.")
            else:
                print("Output path not defined; merged inventory not saved to disk.")
        else:
            print("No merged inventory produced.")
    else:
        print("No inventories were processed.")


if __name__ == "__main__":
    main()